import openpyxl
import json
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET

# Excel dosyasını aç
excel_path = r"C:\Users\ESAT\Desktop\Son\son surec.xlsm"
output_path = r"C:\Users\ESAT\kkp-platform\excel_analysis.json"

print("Excel dosyasi analiz ediliyor...")
print(f"Dosya: {excel_path}\n")

# Excel dosyasını yükle
wb = openpyxl.load_workbook(excel_path, data_only=False, keep_vba=True)

analysis = {
    "dosya_adi": "son surec.xlsm",
    "sayfalar": [],
    "makrolar": {},
    "ozetler": {}
}

# Sayfa bilgilerini topla
print("\nSayfalar:")
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]

    # Chartsheet kontrolü
    if not hasattr(sheet, 'max_row'):
        print(f"  - {sheet_name}: [Grafik Sayfasi]")
        analysis["sayfalar"].append({
            "isim": sheet_name,
            "tip": "chartsheet"
        })
        continue

    # Sayfa boyutları
    max_row = sheet.max_row
    max_col = sheet.max_column

    print(f"  - {sheet_name}: {max_row} satir, {max_col} sutun")

    # İlk birkaç satırı oku (başlıklar için)
    headers = []
    first_rows = []

    for row_idx in range(1, min(6, max_row + 1)):
        row_data = []
        for col_idx in range(1, min(max_col + 1, 20)):  # İlk 20 sütun
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            if value is not None:
                row_data.append({
                    "col": col_idx,
                    "value": str(value)[:100]  # İlk 100 karakter
                })
        if row_data:
            first_rows.append({"row": row_idx, "data": row_data})

    # Sayfa özeti
    page_info = {
        "isim": sheet_name,
        "satir_sayisi": max_row,
        "sutun_sayisi": max_col,
        "ilk_satirlar": first_rows[:5]
    }

    analysis["sayfalar"].append(page_info)

# VBA Makrolarını çıkar
print("\nMakrolar analiz ediliyor...")
try:
    # xlsm dosyası zip olarak açılabilir
    with zipfile.ZipFile(excel_path, 'r') as zip_ref:
        # VBA projesi dosyalarını listele
        vba_files = [f for f in zip_ref.namelist() if 'vba' in f.lower() or 'xl/vba' in f]

        print(f"  VBA dosyaları bulundu: {len(vba_files)}")

        # vbaProject.bin dosyasını ara
        if 'xl/vbaProject.bin' in zip_ref.namelist():
            print("  + vbaProject.bin bulundu - Makrolar mevcut!")
            analysis["makrolar"]["mevcut"] = True
            analysis["makrolar"]["dosya"] = "xl/vbaProject.bin"

            # VBA binary dosyasını oku (text olarak parse edilemez ama var olduğunu kaydediyoruz)
            vba_content = zip_ref.read('xl/vbaProject.bin')
            analysis["makrolar"]["boyut"] = len(vba_content)

            # VBA modül isimlerini bulmaya çalış
            vba_text = str(vba_content)
            if 'Module' in vba_text or 'Sub' in vba_text:
                print("  + Modul ve Sub prosedurler tespit edildi")
                analysis["makrolar"]["icerdigi"] = "Moduller ve Sub prosedurler"

        # customUI ribbon dosyalarını kontrol et
        custom_ui_files = [f for f in zip_ref.namelist() if 'customUI' in f]
        if custom_ui_files:
            print(f"  + Ozel UI dosyalari bulundu: {custom_ui_files}")
            analysis["makrolar"]["custom_ui"] = custom_ui_files

except Exception as e:
    print(f"  ! Makro analizi hatasi: {e}")
    analysis["makrolar"]["hata"] = str(e)

# Adlandırılmış aralıkları kontrol et
print("\nTanimli Isimler (Named Ranges):")
if wb.defined_names:
    analysis["tanimli_isimler"] = []
    for name, defn in wb.defined_names.items():
        print(f"  - {name}: {defn.value}")
        analysis["tanimli_isimler"].append({
            "isim": name,
            "referans": str(defn.value)
        })

# Özet bilgiler
analysis["ozetler"] = {
    "toplam_sayfa": len(wb.sheetnames),
    "makro_var_mi": "xl/vbaProject.bin" in str(analysis.get("makrolar", {})),
    "sayfa_isimleri": wb.sheetnames
}

print(f"\nAnaliz sonuclari kaydediliyor: {output_path}")

# JSON olarak kaydet
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(analysis, f, ensure_ascii=False, indent=2)

print("\nAnaliz tamamlandi!")
print(f"\nOZET:")
print(f"  Toplam Sayfa: {analysis['ozetler']['toplam_sayfa']}")
print(f"  Sayfa Isimleri: {', '.join(analysis['ozetler']['sayfa_isimleri'])}")
print(f"  Makro Var mi: {'Evet' if analysis['ozetler']['makro_var_mi'] else 'Hayir'}")
