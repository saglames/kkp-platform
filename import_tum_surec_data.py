import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import sys
import io

# Fix encoding for Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Excel dosyası
excel_path = r"C:\Users\ESAT\Desktop\Son\son surec.xlsm"

# PostgreSQL bağlantısı
conn = psycopg2.connect(
    host="localhost",
    port=5432,
    database="kkp_db",
    user="postgres",
    password="postgres"
)
cur = conn.cursor()

print("Excel dosyası okunuyor...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Ürün listesini topla
urunler = {}

def process_sheet(sheet_name, adet_column):
    """Sayfadaki urunleri isle"""
    if sheet_name not in wb.sheetnames:
        print(f"  UYARI '{sheet_name}' sayfasi bulunamadi")
        return

    sheet = wb[sheet_name]
    print(f"\n'{sheet_name}' sayfasi isleniyor...")

    # Baslik satirini bul
    header_row = None
    for row_idx in range(1, 10):
        cell_value = sheet.cell(row_idx, 1).value
        if cell_value and ('URUN' in str(cell_value).upper()):
            header_row = row_idx
            break

    if not header_row:
        header_row = 1

    print(f"  Baslik satiri: {header_row}")

    # Veri satırlarını oku
    count = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        urun_kodu = sheet.cell(row_idx, 2).value  # B sütunu: Ürün Kodu
        adet_cell = sheet.cell(row_idx, adet_column).value  # Adet sütunu

        if not urun_kodu or str(urun_kodu).strip() == '':
            continue

        urun_kodu = str(urun_kodu).strip()

        # Adet değerini al
        try:
            adet = int(adet_cell) if adet_cell else 0
        except (ValueError, TypeError):
            adet = 0

        if adet <= 0:
            continue

        # Ürün kodunu kaydet
        if urun_kodu not in urunler:
            urunler[urun_kodu] = {
                'temizlemeye_gidecek': 0,
                'temizlemede_olan': 0,
                'temizlemeden_gelen': 0,
                'sevke_hazir': 0,
                'kalan': 0
            }

        # Sayfaya göre doğru alana ekle
        if 'Temizlemeye Gidecek' in sheet_name or 'GİDECEK' in sheet_name.upper():
            urunler[urun_kodu]['temizlemeye_gidecek'] += adet
        elif 'Temizlemede Olan' in sheet_name or 'OLAN' in sheet_name.upper():
            urunler[urun_kodu]['temizlemede_olan'] += adet
        elif 'Temizlemeden Gelen' in sheet_name or 'GELEN' in sheet_name.upper():
            urunler[urun_kodu]['temizlemeden_gelen'] += adet
        elif 'Sevke Hazır' in sheet_name or 'HAZIR' in sheet_name.upper():
            urunler[urun_kodu]['sevke_hazir'] += adet
        elif 'Kalan' in sheet_name:
            urunler[urun_kodu]['kalan'] += adet

        count += 1

    print(f"  OK {count} urun kaydi islendi")

# Sayfaları işle - adet sütunu her sayfa için 3. sütun (C)
process_sheet('Temizlemeye Gidecek', 3)  # C sütunu: Adet
process_sheet('Temizlemede Olan', 3)     # C sütunu: Adet
process_sheet('Temizlemeden Gelen', 3)   # C sütunu: Adet
process_sheet('Sevke Hazır', 3)          # C sütunu: Adet (toplam)
process_sheet('Kalan', 3)                # C sütunu: Adet

print(f"\n\nToplam {len(urunler)} benzersiz urun bulundu")

# Veritabanina aktar
print("\nVeritabanina aktariliyor...\n")

try:
    # Önce mevcut verileri temizle
    print("Mevcut veriler temizleniyor...")
    cur.execute("TRUNCATE TABLE surec_sevk_edilen, surec_kalan, surec_sevke_hazir, surec_temizlemeden_gelen, surec_temizlemede_olan, surec_temizlemeye_gidecek, surec_hareket_log, surec_urunler RESTART IDENTITY CASCADE")

    # Ürünleri ekle
    for urun_kodu, adetler in urunler.items():
        # Ürün kodundan kalite sınıfını belirle (A veya B)
        tip = 'A'  # Varsayılan
        urun_kodu_base = urun_kodu

        if urun_kodu.endswith('A') or urun_kodu.endswith('-A'):
            tip = 'A'
            urun_kodu_base = urun_kodu.rstrip('A').rstrip('-')
        elif urun_kodu.endswith('B') or urun_kodu.endswith('-B'):
            tip = 'B'
            urun_kodu_base = urun_kodu.rstrip('B').rstrip('-')

        # Ürünü surec_urunler tablosuna ekle
        cur.execute("""
            INSERT INTO surec_urunler (tip, urun_kodu, urun_kodu_base)
            VALUES (%s, %s, %s)
            ON CONFLICT (urun_kodu) DO UPDATE SET
                tip = EXCLUDED.tip,
                urun_kodu_base = EXCLUDED.urun_kodu_base
            RETURNING id
        """, (tip, urun_kodu, urun_kodu_base))

        urun_id = cur.fetchone()[0]

        # Temizlemeye Gidecek
        if adetler['temizlemeye_gidecek'] > 0:
            cur.execute("""
                INSERT INTO surec_temizlemeye_gidecek (urun_id, adet, updated_by)
                VALUES (%s, %s, 'Excel Import')
                ON CONFLICT (urun_id) DO UPDATE SET
                    adet = EXCLUDED.adet,
                    updated_by = EXCLUDED.updated_by
            """, (urun_id, adetler['temizlemeye_gidecek']))

        # Temizlemede Olan
        if adetler['temizlemede_olan'] > 0:
            cur.execute("""
                INSERT INTO surec_temizlemede_olan (urun_id, adet, updated_by)
                VALUES (%s, %s, 'Excel Import')
                ON CONFLICT (urun_id) DO UPDATE SET
                    adet = EXCLUDED.adet,
                    updated_by = EXCLUDED.updated_by
            """, (urun_id, adetler['temizlemede_olan']))

        # Temizlemeden Gelen
        if adetler['temizlemeden_gelen'] > 0:
            cur.execute("""
                INSERT INTO surec_temizlemeden_gelen (urun_id, adet, updated_by)
                VALUES (%s, %s, 'Excel Import')
                ON CONFLICT (urun_id) DO UPDATE SET
                    adet = EXCLUDED.adet,
                    updated_by = EXCLUDED.updated_by
            """, (urun_id, adetler['temizlemeden_gelen']))

        # Sevke Hazır
        if adetler['sevke_hazir'] > 0:
            cur.execute("""
                INSERT INTO surec_sevke_hazir (urun_id, adet, updated_by)
                VALUES (%s, %s, 'Excel Import')
                ON CONFLICT (urun_id) DO UPDATE SET
                    adet = EXCLUDED.adet,
                    updated_by = EXCLUDED.updated_by
            """, (urun_id, adetler['sevke_hazir']))

        # Kalan
        if adetler['kalan'] > 0:
            cur.execute("""
                INSERT INTO surec_kalan (urun_id, adet, updated_by)
                VALUES (%s, %s, 'Excel Import')
                ON CONFLICT (urun_id) DO UPDATE SET
                    adet = EXCLUDED.adet,
                    updated_by = EXCLUDED.updated_by
            """, (urun_id, adetler['kalan']))

        print(f"OK {urun_kodu} ({tip}) - TG:{adetler['temizlemeye_gidecek']}, TO:{adetler['temizlemede_olan']}, TGL:{adetler['temizlemeden_gelen']}, SH:{adetler['sevke_hazir']}, K:{adetler['kalan']}")

    conn.commit()
    print(f"\nBASARILI! {len(urunler)} urun aktarildi!")

except Exception as e:
    conn.rollback()
    print(f"\nHATA: {e}")
    import traceback
    traceback.print_exc()
finally:
    cur.close()
    conn.close()
